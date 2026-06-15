from flask import Blueprint, send_file, request
from flask_login import login_required
from app.models import Producto, Cuadrilla, Salida, SalidaItem, Rendicion, RendicionItem, PrestamoMaquinaria
from app import db
from sqlalchemy import func
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

export_bp = Blueprint('export', __name__, url_prefix='/export')

# Colores ANCO
AZUL       = '3E88D6'
AZUL_OSCURO= '0d1f35'
GRIS_CLARO = 'f0f4f8'
ROJO       = 'e63946'
VERDE      = '2a9d8f'
NARANJO    = 'f4a261'

def estilo_header(ws, fila, cols, color=AZUL_OSCURO):
    for col in range(1, cols + 1):
        cell = ws.cell(row=fila, column=col)
        cell.font      = Font(bold=True, color='FFFFFF', size=10)
        cell.fill      = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center', vertical='center')

def borde_fino():
    lado = Side(style='thin', color='DDDDDD')
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def auto_ancho(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ── STOCK BODEGA ────────────────────────────────────────────────────────────
@export_bp.route('/stock')
@login_required
def stock():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Bodega'

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = f'ANCO BODEGA — Stock al {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_OSCURO)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ['Código', 'Descripción', 'Categoría', 'Unidad', 'Stock Actual', 'Stock Mínimo', 'Estado', 'Diferencia']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', fgColor=AZUL)
        c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    productos = Producto.query.filter_by(activo=True).order_by(Producto.categoria, Producto.descripcion).all()
    for row_n, p in enumerate(productos, 3):
        alerta = p.alerta()
        estado = '⚠️ CRÍTICO' if alerta == 'critico' else ('⚡ BAJO' if alerta == 'bajo' else '✅ Normal')
        dif    = p.stock_bodega - p.stock_min
        row_data = [p.codigo, p.descripcion, p.categoria, p.unidad, p.stock_bodega, p.stock_min, estado, dif]
        for col_n, val in enumerate(row_data, 1):
            c = ws.cell(row=row_n, column=col_n, value=val)
            c.border    = borde_fino()
            c.alignment = Alignment(vertical='center')
            if alerta == 'critico':
                c.fill = PatternFill('solid', fgColor='FFEAEA')
            elif alerta == 'bajo':
                c.fill = PatternFill('solid', fgColor='FFF4E5')
            elif row_n % 2 == 0:
                c.fill = PatternFill('solid', fgColor=GRIS_CLARO)

    auto_ancho(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'ANCO_Stock_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ── LISTA DE COMPRAS ────────────────────────────────────────────────────────
@export_bp.route('/compras')
@login_required
def compras():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lista de Compras'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'ANCO — Lista de Compras al {datetime.now().strftime("%d/%m/%Y")}'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_OSCURO)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    headers = ['Código', 'Descripción', 'Categoría', 'Unidad', 'Stock Actual', 'Cantidad a Reponer']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=ROJO)
        c.alignment = Alignment(horizontal='center')

    productos = Producto.query.filter(
        Producto.activo == True,
        Producto.stock_bodega <= Producto.stock_min
    ).order_by(Producto.stock_bodega).all()

    for row_n, p in enumerate(productos, 3):
        falta = p.stock_min - p.stock_bodega
        row_data = [p.codigo, p.descripcion, p.categoria, p.unidad, p.stock_bodega, max(falta, 1)]
        for col_n, val in enumerate(row_data, 1):
            c = ws.cell(row=row_n, column=col_n, value=val)
            c.border = borde_fino()
            if row_n % 2 == 0:
                c.fill = PatternFill('solid', fgColor='FFF4E5')

    auto_ancho(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'ANCO_Compras_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ── COMPARATIVO MENSUAL ──────────────────────────────────────────────────────
@export_bp.route('/comparativo')
@login_required
def comparativo():
    año  = request.args.get('año',  datetime.now().year, type=int)
    mes  = request.args.get('mes',  datetime.now().month, type=int)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparativo Mensual'

    meses_es = ['', 'Enero','Febrero','Marzo','Abril','Mayo','Junio',
                'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    ws.merge_cells('A1:F1')
    ws['A1'] = f'ANCO — Consumo por Cuadrilla — {meses_es[mes]} {año}'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_OSCURO)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # ── Hoja 1: Resumen por cuadrilla ──
    headers = ['Cuadrilla', 'N° Salidas', 'N° OT Rendidas', 'Productos distintos usados']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=AZUL)
        c.alignment = Alignment(horizontal='center')

    cuadrillas = Cuadrilla.query.filter_by(activa=True).all()
    for row_n, c_obj in enumerate(cuadrillas, 3):
        salidas_mes = Salida.query.filter(
            Salida.cuadrilla_id == c_obj.id,
            db.func.extract('year',  Salida.creado_en) == año,
            db.func.extract('month', Salida.creado_en) == mes
        ).count()
        ots_mes = Rendicion.query.filter(
            Rendicion.cuadrilla_id == c_obj.id,
            db.func.extract('year',  Rendicion.creado_en) == año,
            db.func.extract('month', Rendicion.creado_en) == mes
        ).count()
        prods_distintos = db.session.query(func.count(func.distinct(RendicionItem.producto_id)))\
            .join(Rendicion)\
            .filter(
                Rendicion.cuadrilla_id == c_obj.id,
                db.func.extract('year',  Rendicion.creado_en) == año,
                db.func.extract('month', Rendicion.creado_en) == mes
            ).scalar() or 0

        for col_n, val in enumerate([c_obj.nombre, salidas_mes, ots_mes, prods_distintos], 1):
            cell = ws.cell(row=row_n, column=col_n, value=val)
            cell.border = borde_fino()
            if row_n % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=GRIS_CLARO)

    auto_ancho(ws)

    # ── Hoja 2: Detalle por cuadrilla ──
    ws2 = wb.create_sheet('Detalle Consumo')
    ws2.merge_cells('A1:F1')
    ws2['A1'] = f'Detalle de consumo por OT — {meses_es[mes]} {año}'
    ws2['A1'].font      = Font(bold=True, size=12, color='FFFFFF')
    ws2['A1'].fill      = PatternFill('solid', fgColor=AZUL_OSCURO)
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['Cuadrilla', 'N° OT', 'Fecha', 'Producto', 'Cantidad', 'Unidad']
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=AZUL)
        c.alignment = Alignment(horizontal='center')

    rendiciones = Rendicion.query.filter(
        db.func.extract('year',  Rendicion.creado_en) == año,
        db.func.extract('month', Rendicion.creado_en) == mes
    ).order_by(Rendicion.cuadrilla_id, Rendicion.creado_en).all()

    row_n = 3
    for r in rendiciones:
        for item in r.items:
            for col_n, val in enumerate([
                r.cuadrilla.nombre, r.numero_ot,
                r.creado_en.strftime('%d/%m/%Y'),
                item.producto.descripcion,
                item.cantidad_usada,
                item.producto.unidad
            ], 1):
                cell = ws2.cell(row=row_n, column=col_n, value=val)
                cell.border = borde_fino()
                if row_n % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor=GRIS_CLARO)
            row_n += 1

    auto_ancho(ws2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'ANCO_Comparativo_{meses_es[mes]}_{año}.xlsx')


# ── HISTORIAL SALIDAS ────────────────────────────────────────────────────────
@export_bp.route('/salidas')
@login_required
def salidas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Historial Salidas'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'ANCO — Historial de Salidas al {datetime.now().strftime("%d/%m/%Y")}'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_OSCURO)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Fecha/Hora', 'Cuadrilla', 'Producto', 'Cantidad', 'Unidad', 'Notas']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=AZUL)
        c.alignment = Alignment(horizontal='center')

    salidas_list = Salida.query.order_by(Salida.creado_en.desc()).all()
    row_n = 3
    for s in salidas_list:
        for item in s.items:
            for col_n, val in enumerate([
                s.creado_en.strftime('%d/%m/%Y %H:%M'),
                s.cuadrilla.nombre,
                item.producto.descripcion,
                item.cantidad,
                item.producto.unidad,
                s.notas or ''
            ], 1):
                cell = ws.cell(row=row_n, column=col_n, value=val)
                cell.border = borde_fino()
                if row_n % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor=GRIS_CLARO)
            row_n += 1

    auto_ancho(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'ANCO_Salidas_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ── EXCEL ENTREGADO VS RENDIDO POR CUADRILLA ────────────────────────────────
@export_bp.route('/entregado_rendido')
@login_required
def entregado_rendido():
    from app.models import RendicionSalida, RendicionSalidaItem
    from datetime import date

    desde_str    = request.args.get('desde', '')
    hasta_str    = request.args.get('hasta', '')
    cuadrilla_id = request.args.get('cuadrilla_id', '')

    hoy = date.today()
    desde_dt = datetime.strptime(desde_str, '%Y-%m-%d') if desde_str else datetime(hoy.year, hoy.month, 1)
    hasta_dt = datetime.strptime(hasta_str, '%Y-%m-%d').replace(hour=23, minute=59) if hasta_str else datetime.now()

    cuadrillas = Cuadrilla.query.filter_by(activa=True).all()
    if cuadrilla_id:
        cuadrillas = [c for c in cuadrillas if str(c.id) == str(cuadrilla_id)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for cuadrilla in cuadrillas:
        # Una hoja por cuadrilla
        nombre_hoja = cuadrilla.nombre[:30].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(title=nombre_hoja)

        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f'ANCO — {cuadrilla.nombre} — {desde_dt.strftime("%d/%m/%Y")} al {hasta_dt.strftime("%d/%m/%Y")}'
        ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_OSCURO)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 28

        # ── SECCIÓN ENTREGAS ──
        ws.merge_cells('A2:G2')
        ws['A2'] = '📦 ENTREGAS (SALIDAS DESDE BODEGA)'
        ws['A2'].font = Font(bold=True, size=11, color='FFFFFF')
        ws['A2'].fill = PatternFill('solid', fgColor=AZUL)
        ws['A2'].alignment = Alignment(horizontal='left')
        ws.row_dimensions[2].height = 20

        headers_ent = ['Fecha', 'Producto', 'Código', 'Unidad', 'Cantidad Entregada', 'Notas', 'Registrado por']
        for i, h in enumerate(headers_ent, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = PatternFill('solid', fgColor='2c5282')
            c.alignment = Alignment(horizontal='center')

        salidas = Salida.query.filter(
            Salida.cuadrilla_id == cuadrilla.id,
            Salida.creado_en.between(desde_dt, hasta_dt),
            Salida.anulada == False,
            Salida.tipo == 'salida'
        ).order_by(Salida.creado_en).all()

        row_ent = 4
        total_entregado = 0
        for s in salidas:
            for item in s.items:
                vals = [
                    s.creado_en.strftime('%d/%m/%Y %H:%M'),
                    item.producto.descripcion,
                    item.producto.codigo,
                    item.producto.unidad,
                    item.cantidad,
                    s.notas or '',
                    s.usuario.nombre if s.usuario else '—'
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_ent, column=col, value=val)
                    cell.border = borde_fino()
                    if row_ent % 2 == 0:
                        cell.fill = PatternFill('solid', fgColor=GRIS_CLARO)
                    if col == 5:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                total_entregado += item.cantidad
                row_ent += 1

        if row_ent == 4:
            ws.cell(row=4, column=1, value='Sin entregas en el período').font = Font(italic=True, color='888888')
            row_ent = 5

        # Total entregas
        cell_tot = ws.cell(row=row_ent, column=4, value='TOTAL ENTREGADO:')
        cell_tot.font = Font(bold=True)
        cell_tot.alignment = Alignment(horizontal='right')
        cell_tot2 = ws.cell(row=row_ent, column=5, value=total_entregado)
        cell_tot2.font = Font(bold=True, color=AZUL)
        cell_tot2.fill = PatternFill('solid', fgColor='dbeafe')
        row_ent += 2

        # ── SECCIÓN RENDICIONES ──
        ws.cell(row=row_ent, column=1)
        ws.merge_cells(f'A{row_ent}:G{row_ent}')
        ws[f'A{row_ent}'] = '✅ RENDICIONES (LO UTILIZADO EN TERRENO)'
        ws[f'A{row_ent}'].font = Font(bold=True, size=11, color='FFFFFF')
        ws[f'A{row_ent}'].fill = PatternFill('solid', fgColor=VERDE)
        ws[f'A{row_ent}'].alignment = Alignment(horizontal='left')
        ws.row_dimensions[row_ent].height = 20
        row_ent += 1

        headers_rend = ['Fecha', 'Producto', 'Código', 'Unidad', 'Cantidad Rendida', 'Notas', 'Registrado por']
        for i, h in enumerate(headers_rend, 1):
            c = ws.cell(row=row_ent, column=i, value=h)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = PatternFill('solid', fgColor='1a6b5f')
            c.alignment = Alignment(horizontal='center')
        row_ent += 1

        try:
            rend_salidas = RendicionSalida.query.filter(
                RendicionSalida.cuadrilla_id == cuadrilla.id,
                RendicionSalida.creado_en.between(desde_dt, hasta_dt)
            ).order_by(RendicionSalida.creado_en).all()
        except:
            rend_salidas = []

        total_rendido = 0
        row_rend = row_ent
        for rend in rend_salidas:
            for item in rend.items:
                vals = [
                    rend.creado_en.strftime('%d/%m/%Y %H:%M'),
                    item.producto.descripcion,
                    item.producto.codigo,
                    item.producto.unidad,
                    item.cantidad_rendida,
                    rend.notas or '',
                    rend.usuario.nombre if rend.usuario else '—'
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_rend, column=col, value=val)
                    cell.border = borde_fino()
                    if row_rend % 2 == 0:
                        cell.fill = PatternFill('solid', fgColor='f0fdf8')
                    if col == 5:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                total_rendido += item.cantidad_rendida
                row_rend += 1

        if row_rend == row_ent:
            ws.cell(row=row_ent, column=1, value='Sin rendiciones en el período').font = Font(italic=True, color='888888')
            row_rend = row_ent + 1

        # Total rendido
        cell_tr = ws.cell(row=row_rend, column=4, value='TOTAL RENDIDO:')
        cell_tr.font = Font(bold=True)
        cell_tr.alignment = Alignment(horizontal='right')
        cell_tr2 = ws.cell(row=row_rend, column=5, value=total_rendido)
        cell_tr2.font = Font(bold=True, color=VERDE)
        cell_tr2.fill = PatternFill('solid', fgColor='d1fae5')
        row_rend += 2

        # ── RESUMEN DIFERENCIA ──
        ws.merge_cells(f'A{row_rend}:G{row_rend}')
        ws[f'A{row_rend}'] = '📊 RESUMEN'
        ws[f'A{row_rend}'].font = Font(bold=True, size=11, color='FFFFFF')
        ws[f'A{row_rend}'].fill = PatternFill('solid', fgColor=AZUL_OSCURO)
        ws[f'A{row_rend}'].alignment = Alignment(horizontal='left')
        row_rend += 1

        diferencia = total_rendido - total_entregado
        resumen = [
            ('Total Entregado:', total_entregado, AZUL),
            ('Total Rendido:',   total_rendido,   VERDE),
            ('Diferencia:',      diferencia,      ROJO if diferencia < 0 else VERDE),
        ]
        for label, val, color in resumen:
            ws.cell(row=row_rend, column=3, value=label).font = Font(bold=True)
            ws.cell(row=row_rend, column=3).alignment = Alignment(horizontal='right')
            c = ws.cell(row=row_rend, column=4, value=val)
            c.font = Font(bold=True, color=color, size=12)
            c.alignment = Alignment(horizontal='center')
            row_rend += 1

        auto_ancho(ws)

    # Si no hay cuadrillas con datos crea una hoja de aviso
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet('Sin datos')
        ws['A1'] = 'No hay datos en el período seleccionado'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre = f'ANCO_Entregado_Rendido_{desde_dt.strftime("%d%m%Y")}_{hasta_dt.strftime("%d%m%Y")}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre)


# ── EXCEL COMPARATIVO COMPLETO POR CUADRILLA ────────────────────────────────
@export_bp.route('/comparativo_cuadrilla')
@login_required
def comparativo_cuadrilla():
    from app.models import RendicionSalida, RendicionSalidaItem, Inventario, InventarioItem, StockCuadrilla
    from datetime import date

    desde_str    = request.args.get('desde', '')
    hasta_str    = request.args.get('hasta', '')
    cuadrilla_id = request.args.get('cuadrilla_id', '')
    inventario_id = request.args.get('inventario_id', '')

    hoy = date.today()
    desde_dt = datetime.strptime(desde_str, '%Y-%m-%d') if desde_str else datetime(hoy.year, hoy.month, 1)
    hasta_dt = datetime.strptime(hasta_str, '%Y-%m-%d').replace(hour=23, minute=59) if hasta_str else datetime.now()

    cuadrillas = Cuadrilla.query.filter_by(activa=True).all()
    if cuadrilla_id:
        cuadrillas = [c for c in cuadrillas if str(c.id) == str(cuadrilla_id)]

    # Inventario seleccionado (opcional)
    inventario_sel = None
    if inventario_id:
        inventario_sel = Inventario.query.get(inventario_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    COLOR_HEADER  = 'FFFFFF'
    COL_ENT  = '2c5282'  # azul entregas
    COL_REND = '1a6b5f'  # verde rendiciones salida
    COL_OT   = '7b3f00'  # cafe OT
    COL_INV  = '4a1d6e'  # morado inventario

    for cuadrilla in cuadrillas:
        ws = wb.create_sheet(title=cuadrilla.nombre[:30].replace('/', '-'))

        # ── TÍTULO ──
        ws.merge_cells('A1:H1')
        ws['A1'] = f'ANCO — {cuadrilla.nombre} — {desde_dt.strftime("%d/%m/%Y")} al {hasta_dt.strftime("%d/%m/%Y")}'
        ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor=AZUL_OSCURO)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 28

        if inventario_sel:
            ws.merge_cells('A2:H2')
            ws['A2'] = f'Inventario comparado: {inventario_sel.creado_en.strftime("%d/%m/%Y %H:%M")}'
            ws['A2'].font = Font(bold=True, size=10, color='FFFFFF')
            ws['A2'].fill = PatternFill('solid', fgColor='4a1d6e')
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 18
            fila_header = 3
        else:
            fila_header = 2

        # ── HEADERS ──
        headers = [
            ('Producto',            AZUL_OSCURO),
            ('Código',              AZUL_OSCURO),
            ('Unidad',              AZUL_OSCURO),
            ('Entregado',           COL_ENT),
            ('Rendido (Salidas)',   COL_REND),
            ('Rendido (OT)',        COL_OT),
            ('Total Rendido',       AZUL),
            ('Stock Inventario',    COL_INV),
        ]
        for col, (h, color) in enumerate(headers, 1):
            c = ws.cell(row=fila_header, column=col, value=h)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[fila_header].height = 28

        # ── RECOPILAR DATOS por producto ──
        # Todos los productos que tuvo movimiento en esta cuadrilla
        productos_ids = set()

        # Entregas
        salidas = Salida.query.filter(
            Salida.cuadrilla_id == cuadrilla.id,
            Salida.creado_en.between(desde_dt, hasta_dt),
            Salida.anulada == False,
            Salida.tipo == 'salida'
        ).all()
        for s in salidas:
            for item in s.items:
                productos_ids.add(item.producto_id)

        # Rendiciones por salida
        try:
            rends_salida = RendicionSalida.query.filter(
                RendicionSalida.cuadrilla_id == cuadrilla.id,
                RendicionSalida.creado_en.between(desde_dt, hasta_dt)
            ).all()
            for r in rends_salida:
                for item in r.items:
                    productos_ids.add(item.producto_id)
        except:
            rends_salida = []

        # Rendiciones OT
        rends_ot = Rendicion.query.filter(
            Rendicion.cuadrilla_id == cuadrilla.id,
            Rendicion.creado_en.between(desde_dt, hasta_dt),
            Rendicion.anulada == False
        ).all()
        for r in rends_ot:
            for item in r.items:
                productos_ids.add(item.producto_id)

        # Inventario seleccionado
        inv_dict = {}
        if inventario_sel:
            inv_items = InventarioItem.query.filter_by(
                inventario_id=inventario_sel.id
            ).all()
            for item in inv_items:
                if item.producto.id not in productos_ids:
                    productos_ids.add(item.producto.id)
                inv_dict[item.producto_id] = item.stock_real

        if not productos_ids:
            ws.cell(row=fila_header+1, column=1, value='Sin movimientos en el período').font = Font(italic=True, color='888888')
            auto_ancho(ws)
            continue

        productos = Producto.query.filter(Producto.id.in_(productos_ids)).order_by(Producto.descripcion).all()

        # Calcular totales por producto
        fila = fila_header + 1
        tot_ent, tot_rend_sal, tot_rend_ot, tot_inv = 0, 0, 0, 0

        for p in productos:
            # Entregado
            entregado = sum(
                item.cantidad for s in salidas for item in s.items if item.producto_id == p.id
            )
            # Rendido por salida
            rendido_sal = sum(
                item.cantidad_rendida for r in rends_salida for item in r.items if item.producto_id == p.id
            )
            # Rendido OT
            rendido_ot = sum(
                item.cantidad_usada for r in rends_ot for item in r.items if item.producto_id == p.id
            )
            total_rendido = rendido_sal + rendido_ot
            stock_inv = inv_dict.get(p.id, '—')

            vals = [p.descripcion, p.codigo, p.unidad, entregado, rendido_sal, rendido_ot, total_rendido, stock_inv]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=fila, column=col, value=val)
                cell.border = borde_fino()
                if fila % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor=GRIS_CLARO)
                if col in [4,5,6,7,8]:
                    cell.alignment = Alignment(horizontal='center')
                if col == 4:
                    cell.font = Font(bold=True, color=COL_ENT)
                elif col == 5:
                    cell.font = Font(bold=True, color=COL_REND)
                elif col == 6:
                    cell.font = Font(bold=True, color=COL_OT)
                elif col == 7:
                    cell.font = Font(bold=True, color=AZUL)
                elif col == 8 and stock_inv != '—':
                    cell.font = Font(bold=True, color=COL_INV)

            if isinstance(entregado, (int, float)): tot_ent += entregado
            if isinstance(rendido_sal, (int, float)): tot_rend_sal += rendido_sal
            if isinstance(rendido_ot, (int, float)): tot_rend_ot += rendido_ot
            if isinstance(stock_inv, (int, float)): tot_inv += stock_inv
            fila += 1

        # Fila de totales
        ws.cell(row=fila, column=3, value='TOTALES').font = Font(bold=True)
        ws.cell(row=fila, column=3).alignment = Alignment(horizontal='right')
        for col, val in [(4, tot_ent), (5, tot_rend_sal), (6, tot_rend_ot), (7, tot_rend_sal+tot_rend_ot), (8, tot_inv if tot_inv else '—')]:
            c = ws.cell(row=fila, column=col, value=val)
            c.font = Font(bold=True, size=11)
            c.fill = PatternFill('solid', fgColor='dbeafe')
            c.alignment = Alignment(horizontal='center')
            c.border = borde_fino()

        auto_ancho(ws)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet('Sin datos')
        ws['A1'] = 'No hay datos en el período seleccionado'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = f'ANCO_Comparativo_{desde_dt.strftime("%d%m%Y")}_{hasta_dt.strftime("%d%m%Y")}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=nombre)
